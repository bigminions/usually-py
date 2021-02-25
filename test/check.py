import MySQLdb
import time
import os.path
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook


def tosheet (key, sub2ttles, thetimestart, thetimeend):
    
    
    records = []
    for key in sub2ttles:
        for row in sub2ttles[key]:
            record = []
            record.append(key)
            record.extend(row)
            records.append(record)

    filename = 'check.xlsx'
    if os.path.exists(filename):
        wb = load_workbook(filename = filename)
    else:
        wb = Workbook()

    thetimestart =  '{}'.format(startDay.strftime('%Y-%m-%d'))
    thetimeend =  '{}'.format(endDay.strftime('%Y-%m-%d'))
    title = '{}-{}'.format(thetimestart, thetimeend)
    ws3 = wb.create_sheet(title=title)
    ws3.cell(column=1, row=1, value='gameType')
    ws3.cell(column=2, row=1, value='roomId')
    ws3.cell(column=3, row=1, value='date')
    ws3.cell(column=4, row=1, value='accountId1')
    ws3.cell(column=5, row=1, value='score1')
    ws3.cell(column=6, row=1, value='accountId2')
    ws3.cell(column=7, row=1, value='score2')
    ws3.cell(column=8, row=1, value='accountId3')
    ws3.cell(column=9, row=1, value='score3')
    ws3.cell(column=10, row=1, value='accountId4')
    ws3.cell(column=11, row=1, value='score4')
    for row in range(len(records)):
        for col in range(len(record)):
            ws3.cell(column=col + 1, row=row + 2, value="{0}".format(records[row][col]))
    print "save {}".format(title)
    wb.save(filename)
            
# lastday = 3
# host = 'localhost'
# user = 'root'
# password = ''
# dbname = 'db_game_mj_hz_log'
lastday = 1
host = 'rm-wz9ydv3kqxcu4304k.mysql.rds.aliyuncs.com'
user = 'server_hall'
password = 'Yajing54321'
dbname = 'db_game_mj_hz_log'

db = MySQLdb.connect(host, user, password, dbname, charset='utf8')
cursor = db.cursor()

for i in range(1, lastday + 1):
    now = datetime.now().strftime('%Y-%m-%d %H%M_%S')
    startDay = datetime.now() - timedelta(days=i)
    endDay = datetime.now() - timedelta(days=(i - 1))
    thetimestart =  '{} 00:00:00'.format(startDay.strftime('%Y-%m-%d'))
    thetimeend =  '{} 00:00:00'.format(endDay.strftime('%Y-%m-%d'))
    sql = '''SELECT gameType, roomId, crtTime, accountId1, score1, accountId2, score2, accountId3, score3, accountId4, score4 from t_main_ttlement_log where crtTime > '{}' and crtTime < '{}' and gameType = 'pdk' '''.format(thetimestart, thetimeend);
    print sql
    cursor.execute(sql)
    records = cursor.fetchall()
    roommap = {}
    for record in records:
        gameType = record[0]
        if gameType is not None:
            if gameType not in roommap:
                roommap[gameType] = []
            rooms = roommap[gameType]
            rooms.append(record[1])

    sql = '''SELECT roomId, crtTime, accountId1, score1, accountId2, score2, accountId3, score3, accountId4, score4 from t_sub2_ttlement_log where crtTime > '{}' and crtTime < '{}' '''.format(thetimestart, thetimeend)
    print sql
    cursor.execute(sql)
    records = cursor.fetchall()
    
    sub2ttles = {}
    for key in roommap:
        rooms = roommap[key]
        for record in records:
            if (record[0] in rooms):
                if key not in sub2ttles:
                    sub2ttles[key] = []
                sub2ttles[key].append(record)

    for key in sub2ttles:
        print ('sub2 gametype {} : size = {}'.format(key, len(sub2ttles[key]))) 
                
    tosheet (key, sub2ttles, startDay, endDay)
    
db.close()